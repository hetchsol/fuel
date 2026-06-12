#!/usr/bin/env python3
"""READ-ONLY: inspect the calibration workbook(s) - sheet names, column layout,
row counts, calibrated max, and verify the petrol 49.8cm -> 5071L spot check."""
import openpyxl, sys

def dump(path):
    print("=" * 70); print(path); print("=" * 70)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    print("sheets:", wb.sheetnames)
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        # find header-ish + first/last data rows
        print(f"\n  --- sheet {sn!r} ({ws.max_row} rows x {ws.max_column} cols) ---")
        for r in rows[:4]:
            print("    head:", r)
        # detect dip/vol columns: look for a row layout (blank, dip, vol)
        data = []
        for r in rows:
            # try columns B,C (index1,2) then A,B (0,1)
            for di, vi in ((1, 2), (0, 1)):
                try:
                    d, v = r[di], r[vi]
                except IndexError:
                    continue
                if isinstance(d, (int, float)) and isinstance(v, (int, float)):
                    data.append((float(d), float(v))); break
        if data:
            data.sort()
            print(f"    data pts: {len(data)}  first={data[0]}  last={data[-1]}")
            # petrol spot check near 49.8
            near = [p for p in data if abs(p[0] - 49.8) < 0.06]
            if near: print(f"    @49.8cm -> {near}")

for p in (r"C:\Users\Purchase Requisition\Downloads\DIP CHAT (1) (1).xlsx",
          r"C:\Projects\fuel\DIP CHAT (1).xlsx"):
    try:
        dump(p)
    except Exception as e:
        print("ERR", p, e)
