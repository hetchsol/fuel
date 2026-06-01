"""
Tank Calibration Upload API

Allows owner to upload Excel calibration charts (dip cm → volume L)
per tank. Charts are stored per-station and loaded on startup for
accurate dip-to-volume conversions.
"""

from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from datetime import datetime
from typing import Optional
import io
import logging

from .auth import get_station_context, require_owner, get_current_user
from ...database.station_files import load_station_json, save_station_json
from ...services.audit_service import log_audit_event
from ...services.dip_conversion import register_tank_calibration, _dynamic_calibrations

logger = logging.getLogger(__name__)

router = APIRouter()

CALIBRATION_FILE = 'tank_calibrations.json'


def _load_calibrations(station_id: str) -> dict:
    return load_station_json(station_id, CALIBRATION_FILE, default={})


def _save_calibrations(station_id: str, data: dict):
    save_station_json(station_id, CALIBRATION_FILE, data)


@router.get("/template")
def download_template():
    """Download an Excel template for tank calibration data. No auth required — just a blank template."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Tank Calibration"

    # Instructions
    ws.merge_cells('A1:B1')
    ws['A1'] = "Enter dip stick readings (cm) and corresponding volumes (liters) from the manufacturer's calibration chart"
    ws['A1'].font = Font(italic=True, color="666666")

    # Headers
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col, header in [(1, "Dip (cm)"), (2, "Volume (L)")]:
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Example data (guide)
    examples = [(0, 0), (10, 520), (20, 1850)]
    for i, (dip, vol) in enumerate(examples):
        ws.cell(row=3 + i, column=1, value=dip)
        ws.cell(row=3 + i, column=2, value=vol)

    # Note
    ws.cell(row=7, column=1, value="Replace examples with your actual calibration data")
    ws['A7'].font = Font(italic=True, color="999999")

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="tank_calibration_template.xlsx"'},
    )


@router.post("/upload")
async def upload_calibration(
    file: UploadFile = File(...),
    tank_id: str = Query(...),
    ctx: dict = Depends(get_station_context),
):
    """Upload an Excel calibration chart for a tank (owner only)."""
    try:
        # Inline owner check
        role = ctx.get("role", "")
        role_str = role.value if hasattr(role, 'value') else str(role)
        if role_str != "owner":
            raise HTTPException(status_code=403, detail="Only owners can upload calibration data")

        station_id = ctx["station_id"]
        storage = ctx["storage"]

        # Validate tank exists
        tanks = storage.get('tanks', {})
        if tank_id not in tanks:
            raise HTTPException(status_code=404, detail=f"Tank {tank_id} not found")

        # Validate file type
        if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail=f"File must be .xlsx. Got: {file.filename}")

        # Parse Excel
        from openpyxl import load_workbook
        content = await file.read()
        wb = load_workbook(io.BytesIO(content), data_only=True)

        def _to_num(val):
            """Best-effort numeric coercion. Handles ints/floats, plain numeric
            strings, and comma-decimal locales (e.g. '10,5' -> 10.5)."""
            if val is None:
                return None
            if isinstance(val, (int, float)):
                try:
                    return float(val)
                except (ValueError, TypeError):
                    return None
            if isinstance(val, str):
                s = val.strip().replace(',', '.')
                if not s:
                    return None
                try:
                    return float(s)
                except ValueError:
                    return None
            return None

        def _parse_columns(ws, start_col):
            """Try (start_col, start_col+1) as the (dip, volume) pair."""
            out = {}
            for row in ws.iter_rows(min_col=start_col, max_col=start_col + 1,
                                    values_only=True):
                if not row or len(row) < 2:
                    continue
                dip = _to_num(row[0])
                vol = _to_num(row[1])
                if dip is None or vol is None:
                    continue
                if dip < 0 or vol < 0:
                    continue
                out[dip] = vol
            return out

        def _best_chart_for_sheet(ws):
            """Scan every adjacent column pair on the sheet, pick the densest.

            Lets a sheet that puts data in B/C (or any other adjacent columns)
            still be parsed correctly without forcing the template's A/B layout.
            """
            best: dict = {}
            n_cols = max(ws.max_column or 0, 2)
            for c in range(1, n_cols):
                candidate = _parse_columns(ws, c)
                if len(candidate) > len(best):
                    best = candidate
            return best

        def _norm(s: Optional[str]) -> str:
            import re
            return re.sub(r"[^a-z0-9]", "", (s or "").lower())

        def _sheet_relates_to_tank(sheet_name: str, tid: str) -> bool:
            a, b = _norm(sheet_name), _norm(tid)
            return bool(a) and bool(b) and (a in b or b in a)

        # Two-pass selection so a multi-tank workbook can be uploaded once per
        # tank: first prefer sheets whose normalised name relates to the
        # tank_id (e.g. ' Diesel 2' for TANK-DIESEL-2), then fall back to the
        # densest chart in any sheet.
        chart: dict = {}
        matched_sheet_title: Optional[str] = None
        for ws in wb.worksheets:
            if _sheet_relates_to_tank(ws.title, tank_id):
                candidate = _best_chart_for_sheet(ws)
                if len(candidate) > len(chart):
                    chart = candidate
                    matched_sheet_title = ws.title

        if len(chart) < 5:
            for ws in wb.worksheets:
                candidate = _best_chart_for_sheet(ws)
                if len(candidate) > len(chart):
                    chart = candidate
                    matched_sheet_title = ws.title

        if len(chart) < 5:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Need at least 5 data points (Dip + Volume as two adjacent "
                    f"columns). Found {len(chart)}. Check that your data is in "
                    f"two adjacent columns of a sheet, with numeric values only."
                ),
            )

        # Build sorted chart
        dips = sorted(chart.keys())
        vols = [chart[d] for d in dips]
        sorted_chart = {str(d): chart[d] for d in dips}

        # Get tank capacity
        tank = tanks[tank_id]
        capacity = tank.get("capacity", max(vols) if vols else 50000)

        # Save
        calibrations = _load_calibrations(station_id)
        calibrations[tank_id] = {
            "tank_id": tank_id,
            "chart": sorted_chart,
            "uploaded_at": datetime.now().isoformat(),
            "uploaded_by": ctx.get("username", "unknown"),
            "point_count": len(sorted_chart),
        }
        _save_calibrations(station_id, calibrations)

        # Register for immediate use
        float_chart = {float(k): v for k, v in sorted_chart.items()}
        register_tank_calibration(tank_id, float_chart, capacity=capacity)

        log_audit_event(
            station_id=station_id,
            action="calibration_upload",
            performed_by=ctx.get("username", "unknown"),
            entity_type="tank_calibration",
            entity_id=tank_id,
            details={"point_count": len(sorted_chart)},
        )

        sheet_note = f" from sheet '{matched_sheet_title.strip()}'" if matched_sheet_title else ""
        return {
            "status": "success",
            "message": f"Calibration uploaded: {len(sorted_chart)} data points for {tank_id}{sheet_note}",
            "tank_id": tank_id,
            "point_count": len(sorted_chart),
            "sheet_used": matched_sheet_title.strip() if matched_sheet_title else None,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[calibration] Upload failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


@router.get("/{tank_id}")
def get_calibration(tank_id: str, ctx: dict = Depends(get_station_context)):
    """Get the current calibration chart for a tank."""
    station_id = ctx["station_id"]
    calibrations = _load_calibrations(station_id)

    if tank_id in calibrations:
        return {"found": True, **calibrations[tank_id]}

    return {"found": False, "tank_id": tank_id, "message": "No custom calibration. Using system default."}


@router.delete("/{tank_id}")
def clear_calibration(tank_id: str, ctx: dict = Depends(require_owner)):
    """Clear a tank's custom calibration (owner only). Reverts to system default."""
    station_id = ctx["station_id"]
    calibrations = _load_calibrations(station_id)

    if tank_id not in calibrations:
        raise HTTPException(status_code=404, detail=f"No custom calibration found for {tank_id}")

    del calibrations[tank_id]
    _save_calibrations(station_id, calibrations)

    # Remove from runtime registry
    _dynamic_calibrations.pop(tank_id, None)

    log_audit_event(
        station_id=station_id,
        action="calibration_clear",
        performed_by=ctx["username"],
        entity_type="tank_calibration",
        entity_id=tank_id,
    )

    return {"status": "success", "message": f"Calibration cleared for {tank_id}. Using system default."}


def load_saved_calibrations(station_id: str):
    """Load saved calibrations from JSON and register them. Called at startup."""
    try:
        calibrations = _load_calibrations(station_id)
        for tank_id, cal_data in calibrations.items():
            chart = cal_data.get("chart", {})
            float_chart = {float(k): v for k, v in chart.items()}
            if float_chart:
                register_tank_calibration(tank_id, float_chart)
                logger.info(f"[calibration] Loaded {len(float_chart)} points for {tank_id} at station {station_id}")
    except Exception as e:
        logger.warning(f"[calibration] Failed to load calibrations for station {station_id}: {e}")
