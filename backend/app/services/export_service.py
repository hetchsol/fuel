"""
Export Service — generate CSV and Excel files from station data.
Uses openpyxl (already installed) and stdlib csv/io.
"""
import csv
import io
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------- shared helpers ----------

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header_row(ws, col_count: int, row: int = 1):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _auto_column_widths(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _write_rows(ws, rows: List[list], start_row: int = 2):
    for r_idx, row in enumerate(rows, start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = THIN_BORDER


def _workbook_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(name="Calibri", bold=False, size=10, color="555555")
CENTER_ALIGN = Alignment(horizontal="center")
FOOTER_FONT = Font(name="Calibri", italic=True, size=9, color="999999")


def _add_business_header(ws, col_count: int, business_info: dict = None, report_title: str = ""):
    """Add centered business details header and report title. Returns the next row number for data."""
    from datetime import datetime
    row = 1
    merge_end = get_column_letter(max(col_count, 1))

    if business_info and business_info.get("business_name"):
        # Business name
        ws.merge_cells(f"A{row}:{merge_end}{row}")
        cell = ws.cell(row=row, column=1, value=business_info["business_name"].upper())
        cell.font = TITLE_FONT
        cell.alignment = CENTER_ALIGN
        row += 1

        # Location
        if business_info.get("station_location"):
            ws.merge_cells(f"A{row}:{merge_end}{row}")
            cell = ws.cell(row=row, column=1, value=business_info["station_location"].upper())
            cell.font = SUBTITLE_FONT
            cell.alignment = CENTER_ALIGN
            row += 1

        # Contact
        parts = []
        if business_info.get("contact_phone"): parts.append(business_info["contact_phone"])
        if business_info.get("contact_email"): parts.append(business_info["contact_email"])
        if parts:
            ws.merge_cells(f"A{row}:{merge_end}{row}")
            cell = ws.cell(row=row, column=1, value="  |  ".join(parts))
            cell.font = SUBTITLE_FONT
            cell.alignment = CENTER_ALIGN
            row += 1

        row += 1  # blank row

    # Report title
    if report_title:
        ws.merge_cells(f"A{row}:{merge_end}{row}")
        cell = ws.cell(row=row, column=1, value=report_title.upper())
        cell.font = Font(name="Calibri", bold=True, size=12, color="1F4E79")
        cell.alignment = CENTER_ALIGN
        row += 1

    row += 1  # blank row before data
    return row


def _add_timestamp_footer(ws, row: int, col_count: int):
    """Add timestamp footer below data."""
    from datetime import datetime
    row += 1  # blank row
    merge_end = get_column_letter(max(col_count, 1))
    ws.merge_cells(f"A{row}:{merge_end}{row}")
    cell = ws.cell(row=row, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    cell.font = FOOTER_FONT
    cell.alignment = CENTER_ALIGN


# ---------- Tank Readings ----------

_TR_HEADERS = [
    "Date", "Shift", "Tank", "Fuel Type",
    "Open Dip (cm)", "Close Dip (cm)",
    "Open Vol (L)", "Close Vol (L)", "Tank Movement (L)",
    "Electronic Dispensed (L)", "Mechanical Dispensed (L)",
    "Elec vs Tank Var (L)", "Mech vs Tank Var (L)",
    "Elec vs Tank %", "Mech vs Tank %",
    "Price/L", "Expected Revenue", "Cash Banked", "Cash Diff",
    "Status", "Recorded By",
]


def _tank_reading_row(r: dict) -> list:
    return [
        r.get("date", ""),
        r.get("shift_type", ""),
        r.get("tank_name") or r.get("tank_id", ""),
        r.get("fuel_type", ""),
        r.get("opening_dip_cm"),
        r.get("closing_dip_cm"),
        r.get("opening_volume"),
        r.get("closing_volume"),
        r.get("tank_volume_movement"),
        r.get("total_electronic_dispensed"),
        r.get("total_mechanical_dispensed"),
        r.get("electronic_vs_tank_variance"),
        r.get("mechanical_vs_tank_variance"),
        _pct(r.get("electronic_vs_tank_percent")),
        _pct(r.get("mechanical_vs_tank_percent")),
        r.get("price_per_liter"),
        r.get("expected_amount_electronic"),
        r.get("actual_cash_banked"),
        r.get("cash_difference"),
        r.get("validation_status", ""),
        r.get("recorded_by", ""),
    ]


def _pct(v):
    if v is None:
        return None
    return round(v, 2)


def tank_readings_to_csv(readings: List[dict]) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(_TR_HEADERS)
    for r in readings:
        writer.writerow(_tank_reading_row(r))
    return buf.getvalue()


def tank_readings_to_excel(readings: List[dict], station_name: str = "Station", business_info: dict = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tank Readings"

    col_count = len(_TR_HEADERS)
    data_start = _add_business_header(ws, col_count, business_info, "Tank Readings Report")

    # Header row
    for i, h in enumerate(_TR_HEADERS, 1):
        ws.cell(row=data_start, column=i, value=h)
    _style_header_row(ws, col_count, data_start)

    rows = [_tank_reading_row(r) for r in readings]
    _write_rows(ws, rows, start_row=data_start + 1)
    _auto_column_widths(ws)
    _add_timestamp_footer(ws, data_start + len(rows), col_count)

    ws.sheet_properties.tabColor = "1F4E79"
    return _workbook_to_bytes(wb)


# ---------- Sales ----------

_SALES_HEADERS = [
    "Sale ID", "Shift", "Date", "Fuel Type",
    "Mech Open", "Mech Close", "Mech Volume",
    "Elec Open", "Elec Close", "Elec Volume",
    "Discrepancy %", "Avg Volume", "Unit Price", "Total Amount",
    "Status", "Message",
]


def _sale_row(s: dict) -> list:
    return [
        s.get("sale_id", ""),
        s.get("shift_id", ""),
        s.get("date", ""),
        s.get("fuel_type", ""),
        s.get("mechanical_opening"),
        s.get("mechanical_closing"),
        s.get("mechanical_volume"),
        s.get("electronic_opening"),
        s.get("electronic_closing"),
        s.get("electronic_volume"),
        _pct(s.get("discrepancy_percent")),
        s.get("average_volume"),
        s.get("unit_price"),
        s.get("total_amount"),
        s.get("validation_status", ""),
        s.get("validation_message", ""),
    ]


def sales_to_csv(sales: List[dict]) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(_SALES_HEADERS)
    for s in sales:
        writer.writerow(_sale_row(s))
    return buf.getvalue()


def sales_to_excel(sales: List[dict], station_name: str = "Station", business_info: dict = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"

    col_count = len(_SALES_HEADERS)
    data_start = _add_business_header(ws, col_count, business_info, "Sales Report")

    for i, h in enumerate(_SALES_HEADERS, 1):
        ws.cell(row=data_start, column=i, value=h)
    _style_header_row(ws, col_count, data_start)

    rows = [_sale_row(s) for s in sales]
    _write_rows(ws, rows, start_row=data_start + 1)
    _auto_column_widths(ws)
    _add_timestamp_footer(ws, data_start + len(rows), col_count)

    ws.sheet_properties.tabColor = "2E7D32"
    return _workbook_to_bytes(wb)


# ---------- Reconciliation ----------

_RECON_HEADERS = [
    "Shift ID", "Date", "Shift Type",
    "Petrol Revenue", "Diesel Revenue",
    "LPG Revenue", "Lubricants Revenue", "Accessories Revenue",
    "Total Expected", "Credit Sales", "Expected Cash",
    "Actual Deposited", "Difference", "Cumulative Diff",
    "Notes",
]


def _recon_row(r: dict) -> list:
    return [
        r.get("shift_id", ""),
        r.get("date", ""),
        r.get("shift_type", ""),
        r.get("petrol_revenue"),
        r.get("diesel_revenue"),
        r.get("lpg_revenue"),
        r.get("lubricants_revenue"),
        r.get("accessories_revenue"),
        r.get("total_expected"),
        r.get("credit_sales_total"),
        r.get("expected_cash"),
        r.get("actual_deposited"),
        r.get("difference"),
        r.get("cumulative_difference"),
        r.get("notes", ""),
    ]


def reconciliation_to_csv(recon_data: List[dict]) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(_RECON_HEADERS)
    for r in recon_data:
        writer.writerow(_recon_row(r))
    return buf.getvalue()


def reconciliation_to_excel(recon_data: List[dict], station_name: str = "Station", business_info: dict = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reconciliation"

    col_count = len(_RECON_HEADERS)
    data_start = _add_business_header(ws, col_count, business_info, "Reconciliation Report")

    for i, h in enumerate(_RECON_HEADERS, 1):
        ws.cell(row=data_start, column=i, value=h)
    _style_header_row(ws, col_count, data_start)

    rows = [_recon_row(r) for r in recon_data]
    _write_rows(ws, rows, start_row=data_start + 1)
    _auto_column_widths(ws)
    _add_timestamp_footer(ws, data_start + len(rows), col_count)

    ws.sheet_properties.tabColor = "C62828"
    return _workbook_to_bytes(wb)
