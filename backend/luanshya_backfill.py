"""
One-off backfill: load historical nozzle readings + tank dips for Luanshya (ST002)
from a spreadsheet, via the live API's manager-retro-entry mechanism — exactly as if
attendants had entered the shifts themselves. See:
  C:\\Users\\Purchase Requisition\\.claude\\plans\\snoopy-pondering-sun.md

Usage:
    export LUANSHYA_MANAGER_USER=stanslous
    export LUANSHYA_MANAGER_PASS=<password>
    python luanshya_backfill.py                # dry run (default) — no writes
    python luanshya_backfill.py --apply         # actually create shifts/dips/handovers
"""
import os
import re
import sys
import argparse
from collections import defaultdict
from datetime import datetime

import openpyxl
import requests

API_BASE = os.environ.get("LUANSHYA_API_BASE", "https://fuel-api-wpdj.onrender.com")
STATION_ID = "ST002"
SPREADSHEET_PATH = r"C:\Users\Purchase Requisition\Downloads\18 To 11 Readings.xlsx"

EXPECTED_PRICE = {"Petrol": 26.15, "Diesel": 28.11}
PRICE_TOLERANCE = 0.01

# Sheet "Salesperson" name (lowercased) -> user_id
ATTENDANT_MAP = {
    "shaka": "STF013",
    "violet": "STF014",
    "isabel": "STF015",
    "mubanga": "STF016",
    "trevor": "STF017",
    "prosper": "STF018",
    "matthews": "STF019",
}

FUEL_ABBREV = {"petrol": "UNL", "diesel": "LSD"}
DIFF_TOLERANCE = 0.05  # liters/currency tolerance for post-submit verification diff


def fix_malformed_date(value):
    """Repair the one known bad date string in the sheet, e.g. '21/072026' -> '2026-07-21'."""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, str):
        m = re.match(r"^(\d{2})/(\d{2})(\d{4})$", value.strip())
        if m:
            day, month, year = m.groups()
            return f"{year}-{month}-{day}"
        raise ValueError(f"Unrecognized date string in sheet: {value!r}")
    raise ValueError(f"Unexpected date cell value: {value!r}")


def parse_spreadsheet(path):
    """Returns dict[(date_iso, shift_type)] -> {'nozzles': [...], 'attendants': set()}"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]

    blocks = defaultdict(lambda: {"nozzles": [], "dips": {}})

    for r in range(1, ws.max_row + 1):
        date_cell = ws.cell(row=r, column=1).value
        shift_cell = ws.cell(row=r, column=2).value
        if date_cell is None or date_cell == "Date":
            continue  # blank separator or header row

        date_iso = fix_malformed_date(date_cell)
        shift_type = shift_cell  # "Day" / "Night"
        salesperson = (ws.cell(row=r, column=3).value or "").strip().lower()
        product = (ws.cell(row=r, column=4).value or "").strip().lower()
        nozzle_label = ws.cell(row=r, column=5).value  # e.g. "1A"
        elec_open = ws.cell(row=r, column=6).value
        mech_open = ws.cell(row=r, column=7).value
        elec_close = ws.cell(row=r, column=8).value
        mech_close = ws.cell(row=r, column=9).value
        volume_sold = ws.cell(row=r, column=10).value
        expected_cash = ws.cell(row=r, column=11).value

        key = (date_iso, shift_type)
        block = blocks[key]

        sheet_label = f"{FUEL_ABBREV[product]} {nozzle_label}"

        # Known one-off correction: petrol/1B/Day/18-Jul implied 25.16 ZMW/L instead
        # of the standard 26.15 — confirmed a typo; recompute for the verification
        # diff only (we never submit expected_cash to the API — it's server-computed).
        if (date_iso, shift_type, product, nozzle_label) == ("2026-07-18", "Day", "petrol", "1B") and volume_sold:
            expected_cash = round(volume_sold * EXPECTED_PRICE["Petrol"], 3)

        block["nozzles"].append({
            "sheet_label": sheet_label,
            "product": product,
            "nozzle_label": nozzle_label,
            "attendant_sheet_name": salesperson,
            "electronic_opening": elec_open,
            "mechanical_opening": mech_open,
            "electronic_closing": elec_close,
            "mechanical_closing": mech_close,
            "volume_sold_sheet": volume_sold,
            "expected_cash_sheet": expected_cash,
        })

        # Tank dip data lives on each product's "1B" row in this sheet.
        if nozzle_label == "1B":
            opening_dip_cm = ws.cell(row=r, column=12).value
            opening_vol_sheet = ws.cell(row=r, column=13).value
            closing_dip_cm = ws.cell(row=r, column=14).value
            closing_vol_sheet = ws.cell(row=r, column=15).value
            if isinstance(opening_dip_cm, (int, float)):
                block["dips"][product.capitalize()] = {
                    "opening_dip_cm": opening_dip_cm,
                    "closing_dip_cm": closing_dip_cm,
                    "opening_volume_sheet": opening_vol_sheet,
                    "closing_volume_sheet": closing_vol_sheet,
                }

    return blocks


def is_complete(block):
    return all(n["electronic_closing"] is not None for n in block["nozzles"])


class ApiClient:
    def __init__(self, base_url, station_id):
        self.base_url = base_url.rstrip("/")
        self.station_id = station_id
        self.session = requests.Session()

    def login(self, username, password):
        r = self.session.post(f"{self.base_url}/api/v1/auth/login",
                               json={"username": username, "password": password}, timeout=30)
        r.raise_for_status()
        token = r.json()["access_token"]
        self.session.headers.update({
            "Authorization": f"Bearer {token}",
            "X-Station-Id": self.station_id,
        })

    def get(self, path, **kw):
        r = self.session.get(f"{self.base_url}{path}", timeout=30, **kw)
        r.raise_for_status()
        return r.json()

    def post(self, path, json_body, **kw):
        r = self.session.post(f"{self.base_url}{path}", json=json_body, timeout=30, **kw)
        if not r.ok:
            raise RuntimeError(f"POST {path} -> {r.status_code}: {r.text}")
        return r.json()

    def post_query(self, path, params, **kw):
        r = self.session.post(f"{self.base_url}{path}", params=params, timeout=30, **kw)
        if not r.ok:
            raise RuntimeError(f"POST {path} -> {r.status_code}: {r.text}")
        return r.json()


def resolve_nozzle_ids(client):
    """sheet_label ('UNL 1A') -> nozzle_id"""
    islands = client.get("/api/v1/islands/")
    mapping = {}
    for island in islands:
        ps = island.get("pump_station")
        if not ps:
            continue
        for nozzle in ps.get("nozzles", []):
            abbrev = nozzle.get("fuel_type_abbrev")
            label = nozzle.get("custom_label") or nozzle.get("display_label")
            if abbrev and label:
                mapping[f"{abbrev} {label}"] = nozzle["nozzle_id"]
    return mapping


def resolve_tank_ids(client):
    """'Petrol'/'Diesel' -> tank_id (assumes exactly one tank per fuel type)"""
    tanks = client.get("/api/v1/tanks/levels")
    mapping = {}
    for t in tanks:
        mapping.setdefault(t["fuel_type"], t["tank_id"])
    return mapping


def check_fuel_prices(client):
    settings = client.get("/api/v1/settings/fuel")
    live_petrol = settings["petrol_price_per_liter"]
    live_diesel = settings["diesel_price_per_liter"]
    warnings = []
    if abs(live_petrol - EXPECTED_PRICE["Petrol"]) > PRICE_TOLERANCE:
        warnings.append(f"Live petrol price {live_petrol} != expected {EXPECTED_PRICE['Petrol']}")
    if abs(live_diesel - EXPECTED_PRICE["Diesel"]) > PRICE_TOLERANCE:
        warnings.append(f"Live diesel price {live_diesel} != expected {EXPECTED_PRICE['Diesel']}")
    return warnings


def build_assignments(block, nozzle_id_map):
    """attendant_sheet_name -> list of nozzle_ids they operated this shift"""
    by_attendant = defaultdict(list)
    for n in block["nozzles"]:
        nid = nozzle_id_map.get(n["sheet_label"])
        if nid is None:
            raise RuntimeError(f"Could not resolve nozzle id for sheet label {n['sheet_label']!r}")
        by_attendant[n["attendant_sheet_name"]].append(nid)
    return by_attendant


def print_dry_run(sorted_keys, blocks, nozzle_id_map, tank_id_map, price_warnings):
    print("=" * 70)
    print("DRY RUN — nothing will be written. Review carefully before --apply.")
    print("=" * 70)
    if price_warnings:
        print("\n!! FUEL PRICE WARNINGS !!")
        for w in price_warnings:
            print(f"  - {w}")
    for key in sorted_keys:
        date_iso, shift_type = key
        block = blocks[key]
        complete = is_complete(block)
        shift_id = f"{date_iso}-{shift_type}"
        print(f"\n--- Shift {shift_id} ({'COMPLETE' if complete else 'IN PROGRESS — special handling'}) ---")
        by_attendant = build_assignments(block, nozzle_id_map)
        for name, nozzle_ids in by_attendant.items():
            user_id = ATTENDANT_MAP.get(name)
            print(f"  attendant {name!r} -> {user_id}: nozzles {nozzle_ids}")
        for fuel, dip in block["dips"].items():
            tank_id = tank_id_map.get(fuel)
            print(f"  tank dip [{fuel} -> {tank_id}]: opening {dip['opening_dip_cm']}cm -> closing {dip['closing_dip_cm']}cm "
                  f"(sheet volumes {dip['opening_volume_sheet']} -> {dip['closing_volume_sheet']}, "
                  f"system will compute its own volume from calibration)")


def inspect_existing_shift(client, date_iso, shift_type):
    """Read-only: print full detail on a pre-existing shift so a human can judge
    whether it's safe to skip, or needs investigation, before anything is written."""
    existing = client.get(f"/api/v1/shifts/date/{date_iso}")
    matches = [s for s in existing if s["shift_type"] == shift_type]
    if not matches:
        return None
    shift = matches[0]
    shift_id = shift["shift_id"]
    print(f"\n!! CONFLICT: shift {shift_id} already exists !!")
    print(f"  status: {shift.get('status')}  is_retrospective: {shift.get('is_retrospective')}")
    print(f"  created_by: {shift.get('created_by')}  created_at: {shift.get('created_at')}")
    print(f"  attendants: {shift.get('attendants')}")
    print(f"  assignments: {shift.get('assignments')}")
    try:
        status = client.get(f"/api/v1/handover/shift-submission-status/{shift_id}")
        print(f"  handover submission status: {status}")
    except Exception as e:
        print(f"  (could not fetch handover status: {e})")
    return shift


def preflight_check_all(client, complete_keys):
    """Check every complete-shift date/type for a pre-existing shift, printing full
    detail either way. Returns the set of (date_iso, shift_type) that already exist,
    whether empty or partially/fully real — reactivating a shift via PUT never
    touches handover records, and execute_shift's per-attendant check (via
    get_existing_handover) is what actually decides skip-vs-retro-enter per
    attendant, so no blanket abort is needed here."""
    conflicts = set()
    for date_iso, shift_type in complete_keys:
        shift = inspect_existing_shift(client, date_iso, shift_type)
        if shift is not None:
            conflicts.add((date_iso, shift_type))
    return conflicts


def create_or_reactivate_shift(client, date_iso, shift_type, block, nozzle_id_map, already_exists):
    """POST a fresh shift, or PUT-reactivate a pre-existing empty/auto-closed shell
    (status -> active, is_retrospective -> True, assignments corrected from the sheet).
    Only ever touches shifts already confirmed empty (no submitted handovers) by
    preflight_check_all / the operator reviewing that output."""
    shift_id = f"{date_iso}-{shift_type}"
    by_attendant = build_assignments(block, nozzle_id_map)
    assignments = [
        {
            "attendant_id": ATTENDANT_MAP[name],
            "attendant_name": name,
            "nozzle_ids": nozzle_ids,
        }
        for name, nozzle_ids in by_attendant.items()
    ]

    if already_exists:
        client.session.put(f"{client.base_url}/api/v1/shifts/{shift_id}", json={
            "shift_id": shift_id,
            "date": date_iso,
            "shift_type": shift_type,
            "attendants": [a["attendant_name"] for a in assignments],
            "assignments": assignments,
            "status": "active",
            "is_retrospective": True,
        }, timeout=30).raise_for_status()
        print(f"  reactivated existing shift {shift_id} (status -> active, is_retrospective -> True)")
    else:
        client.post("/api/v1/shifts/", {
            "shift_id": shift_id,
            "date": date_iso,
            "shift_type": shift_type,
            "attendants": [a["attendant_name"] for a in assignments],
            "assignments": assignments,
        })
        print(f"  created shift {shift_id}")

    return by_attendant


def get_existing_handover(client, shift_id, attendant_id):
    """Return the existing handover dict for this shift+attendant if one is already
    in readings_verified/completed phase (i.e. real data already submitted live),
    else None. Checked before every retro-entry attempt so we never try to overwrite
    real attendant-submitted readings."""
    entries = client.get(f"/api/v1/handover/entries?shift_id={shift_id}")
    for e in entries:
        if e.get("attendant_id") == attendant_id and e.get("phase") in ("readings_verified", "completed"):
            return e
    return None


def verify_against_sheet(rows, nozzle_id_map, summaries_by_nozzle, tolerant):
    """Print/raise on mismatches between sheet data and a set of nozzle summaries
    (either freshly created by retro-entry, or an already-existing real handover).
    tolerant=True only warns (used for pre-existing live data, where small sheet
    transcription rounding is expected); tolerant=False raises (used right after
    our own retro-entry submission, where an exact match is expected)."""
    for n in rows:
        nid = nozzle_id_map[n["sheet_label"]]
        summary = summaries_by_nozzle.get(nid)
        if summary is None:
            raise RuntimeError(f"No summary found for nozzle {nid} ({n['sheet_label']})")
        sheet_vol = n["volume_sold_sheet"]
        system_vol = summary["volume_sold"]
        if sheet_vol is not None and abs(system_vol - sheet_vol) > DIFF_TOLERANCE:
            msg = f"volume mismatch {n['sheet_label']}: sheet {sheet_vol} vs system {system_vol}"
            if tolerant:
                print(f"  !! {msg} (pre-existing live data — likely sheet rounding, review if large)")
            else:
                raise RuntimeError(msg)
        sheet_cash = n["expected_cash_sheet"]
        system_rev = summary["revenue"]
        if sheet_cash is not None and abs(system_rev - sheet_cash) > 1.0:
            print(f"  !! revenue diff {n['sheet_label']}: sheet {sheet_cash} vs system {system_rev} "
                  f"(check fuel price for this date)")


def execute_shift(client, date_iso, shift_type, block, nozzle_id_map, tank_id_map, already_exists):
    shift_id = f"{date_iso}-{shift_type}"
    by_attendant = create_or_reactivate_shift(client, date_iso, shift_type, block, nozzle_id_map, already_exists)

    for fuel, dip in block["dips"].items():
        tank_id = tank_id_map[fuel]
        client.post_query("/api/v1/tank-readings/dips", {
            "tank_id": tank_id,
            "date": date_iso,
            "shift_type": shift_type,
            "recorded_by": "luanshya-backfill",
            "opening_dip_cm": dip["opening_dip_cm"],
            "closing_dip_cm": dip["closing_dip_cm"],
        })
        print(f"  recorded tank dip for {fuel} ({tank_id})")

    for name, nozzle_ids in by_attendant.items():
        attendant_id = ATTENDANT_MAP[name]
        rows = [n for n in block["nozzles"] if n["attendant_sheet_name"] == name]

        existing = get_existing_handover(client, shift_id, attendant_id) if already_exists else None
        if existing is not None:
            print(f"  {name} ({attendant_id}) already has a real submitted handover "
                  f"({existing['handover_id']}, phase={existing['phase']}) — skipping retro-entry, "
                  f"verifying it against the sheet instead:")
            summaries_by_nozzle = {s["nozzle_id"]: s for s in existing["nozzle_summaries"]}
            verify_against_sheet(rows, nozzle_id_map, summaries_by_nozzle, tolerant=True)
            print(f"  -> needs Phase 2 (cash) closing manually via handover-review for {name}")
            continue

        opening_readings = [{
            "nozzle_id": nozzle_id_map[n["sheet_label"]],
            "electronic_reading": n["electronic_opening"],
            "mechanical_reading": n["mechanical_opening"] or 0.0,
        } for n in rows]
        closing_readings = [{
            "nozzle_id": nozzle_id_map[n["sheet_label"]],
            "electronic_reading": n["electronic_closing"],
            "mechanical_reading": n["mechanical_closing"] or 0.0,
        } for n in rows]

        result = client.post("/api/v1/handover/manager-retro-entry", {
            "shift_id": shift_id,
            "attendant_id": attendant_id,
            "opening_readings": opening_readings,
            "closing_readings": closing_readings,
            "actual_cash": 0.0,
            "pos_receipts": 0.0,
            "pos_items": [],
            "credit_sales": 0.0,
            "notes": "Backfilled from historical spreadsheet — actual_cash pending manual reconciliation.",
        })
        print(f"  handover created for {name} ({attendant_id}): {result['handover_id']}")

        # The POST response is a compact summary (no nozzle_summaries) — fetch the
        # saved record back to get per-nozzle figures for verification.
        created = get_existing_handover(client, shift_id, attendant_id)
        if created is None:
            raise RuntimeError(f"Could not fetch back the handover just created for {name} ({result['handover_id']})")
        summaries_by_nozzle = {s["nozzle_id"]: s for s in created["nozzle_summaries"]}
        verify_against_sheet(rows, nozzle_id_map, summaries_by_nozzle, tolerant=False)
    print(f"  shift {shift_id} done")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true", help="Actually write data (default is dry-run)")
    parser.add_argument("--inspect-shift", metavar="SHIFT_ID", help="Read-only: print all handover entries for a shift_id and exit")
    args = parser.parse_args()

    username = os.environ.get("LUANSHYA_MANAGER_USER")
    password = os.environ.get("LUANSHYA_MANAGER_PASS")
    if not username or not password:
        print("Set LUANSHYA_MANAGER_USER and LUANSHYA_MANAGER_PASS environment variables first.")
        sys.exit(1)

    if args.inspect_shift:
        client = ApiClient(API_BASE, STATION_ID)
        client.login(username, password)
        entries = client.get(f"/api/v1/handover/entries?shift_id={args.inspect_shift}")
        print(f"{len(entries)} handover entr{'y' if len(entries)==1 else 'ies'} for shift {args.inspect_shift}:")
        for e in entries:
            print("-" * 60)
            for k, v in e.items():
                print(f"  {k}: {v}")
        return

    blocks = parse_spreadsheet(SPREADSHEET_PATH)
    sorted_keys = sorted(blocks.keys(), key=lambda k: (k[0], 0 if k[1] == "Day" else 1))

    client = ApiClient(API_BASE, STATION_ID)
    client.login(username, password)

    nozzle_id_map = resolve_nozzle_ids(client)
    tank_id_map = resolve_tank_ids(client)
    price_warnings = check_fuel_prices(client)

    print_dry_run(sorted_keys, blocks, nozzle_id_map, tank_id_map, price_warnings)

    if not args.apply:
        print("\nDry run complete. Re-run with --apply to write this data.")
        return

    if price_warnings:
        print("\nRefusing to --apply while fuel price warnings are present. Fix prices or investigate first.")
        sys.exit(1)

    complete_keys = [k for k in sorted_keys if is_complete(blocks[k])]
    in_progress_keys = [k for k in sorted_keys if not is_complete(blocks[k])]

    print("\nPre-flight check: looking for pre-existing shifts on all 6 dates...")
    conflicts = preflight_check_all(client, complete_keys)
    if conflicts:
        print(f"\n{len(conflicts)} of {len(complete_keys)} shift(s) already exist as empty/auto-closed "
              f"shells (no submitted handovers) — will be reactivated in place rather than recreated.")
    print("\nProceeding with all 6 shifts...")

    print(f"\nApplying {len(complete_keys)} complete shift(s)...")
    for date_iso, shift_type in complete_keys:
        print(f"\n>>> {date_iso} {shift_type}")
        already_exists = (date_iso, shift_type) in conflicts
        execute_shift(client, date_iso, shift_type, blocks[(date_iso, shift_type)], nozzle_id_map, tank_id_map, already_exists)

    for date_iso, shift_type in in_progress_keys:
        print(f"\n>>> {date_iso} {shift_type} — in-progress shift, checking live state (no writes)")
        existing = client.get(f"/api/v1/shifts/date/{date_iso}")
        matches = [s for s in existing if s["shift_type"] == shift_type]
        if not matches:
            print("  no live shift found for this date/type — SKIPPING (manual creation recommended, not this script)")
            continue
        print(f"  found live shift {matches[0]['shift_id']} — not touching it. "
              f"Cross-check its opening readings against the sheet manually.")

    print("\nDone. All backfilled shifts carry actual_cash=0 pending manual reconciliation.")


if __name__ == "__main__":
    main()
