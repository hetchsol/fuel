"""
Tests for the tank-calibration upload parser: must tolerate real-world
spreadsheet variants (data on a non-active sheet, comma decimals, header rows)
and surface a clear "Found N" message when there genuinely isn't enough data.
"""
import io

import pytest
from openpyxl import Workbook

from app.database import stations_registry


def _xlsx_bytes(builder) -> bytes:
    """Return a workbook (built by `builder(wb)`) serialised to bytes."""
    wb = Workbook()
    wb.remove(wb.active)
    builder(wb)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@pytest.fixture
def tank_st001():
    """Seed a tank directly into the ST001 storage so calibration uploads work."""
    from app.database.storage import get_station_storage
    stations_registry.STATIONS.setdefault("ST001", {
        "station_id": "ST001", "name": "Default", "status": "active",
    })
    storage = get_station_storage("ST001")
    storage.setdefault("tanks", {})["TANK-CAL-1"] = {
        "tank_id": "TANK-CAL-1", "fuel_type": "Diesel",
    }
    return "TANK-CAL-1"


def _upload(client, owner_headers, tank_id, content: bytes):
    return client.post(
        f"/api/v1/settings/tank-calibration/upload?tank_id={tank_id}",
        headers={k: v for k, v in owner_headers.items() if k != "Content-Type"},
        files={"file": ("upload.xlsx", content,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )


def test_data_on_non_active_sheet_still_parses(client, owner_headers, tank_st001):
    """Picks the sheet with the most dip/volume pairs (was a 'Found 0' case)."""
    def build(wb):
        info = wb.create_sheet("Info")
        info["A1"] = "Tank manufacturer notes go here."
        data = wb.create_sheet("Chart")
        data.append(["Dip (cm)", "Volume (L)"])
        for d, v in [(0, 0), (5, 250), (10, 520), (15, 1100), (20, 1850), (25, 2700)]:
            data.append([d, v])
    res = _upload(client, owner_headers, tank_st001, _xlsx_bytes(build))
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["point_count"] >= 5
    assert "Chart" in (body.get("sheet_used") or "") or body["point_count"] >= 5


def test_comma_decimal_locale_parses(client, owner_headers, tank_st001):
    """European 10,5 must coerce to 10.5 rather than being silently dropped."""
    def build(wb):
        ws = wb.create_sheet("Calibration")
        ws.append(["Dip (cm)", "Volume (L)"])
        for d, v in [("0", "0"), ("5,5", "250,2"), ("10,5", "520,1"),
                     ("15,5", "1100,7"), ("20,5", "1850,3"), ("25,5", "2700,8")]:
            ws.append([d, v])
    res = _upload(client, owner_headers, tank_st001, _xlsx_bytes(build))
    assert res.status_code == 200, res.text


def test_too_few_points_gives_clear_error(client, owner_headers, tank_st001):
    def build(wb):
        ws = wb.create_sheet("Chart")
        ws.append(["Dip (cm)", "Volume (L)"])
        ws.append([0, 0])
        ws.append([10, 520])
    res = _upload(client, owner_headers, tank_st001, _xlsx_bytes(build))
    assert res.status_code == 400
    detail = res.json()["detail"]
    assert "Found 2" in detail
    assert "two adjacent columns" in detail


def test_columns_b_c_layout_parses(client, owner_headers, tank_st001):
    """User's real file puts Dip in B and Volume in C (column A empty).
    Parser must auto-detect the columns rather than insisting on A/B."""
    def build(wb):
        ws = wb.create_sheet("Diesel 2")
        # Column A intentionally empty; data in B and C with a header row.
        ws["B2"] = "Dip"
        for i, (d, v) in enumerate(
            [(0, 0), (5, 133), (10, 374), (15, 681), (20, 1041), (25, 1443), (30, 1881)],
            start=3,
        ):
            ws.cell(row=i, column=2, value=d)
            ws.cell(row=i, column=3, value=v)
    res = _upload(client, owner_headers, tank_st001, _xlsx_bytes(build))
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["point_count"] >= 5
    assert body["sheet_used"] == "Diesel 2"


def test_multi_tank_workbook_picks_sheet_matching_tank_id(client, owner_headers):
    """One workbook with three tank sheets — uploading with tank_id=TANK-PETROL-1
    should pick the Petrol 1 sheet (not whichever sheet has the most points)."""
    from app.database.storage import get_station_storage
    storage = get_station_storage("ST001")
    storage.setdefault("tanks", {})["TANK-PETROL-1"] = {
        "tank_id": "TANK-PETROL-1", "fuel_type": "Petrol",
    }

    def build(wb):
        diesel2 = wb.create_sheet(" Diesel 2")
        diesel2.append(["Dip (cm)", "Volume (L)"])
        for d, v in [(0, 0), (5, 100), (10, 220), (15, 360), (20, 520), (25, 700)]:
            diesel2.append([d, v])

        # Larger Petrol 1 sheet, B/C layout.
        petrol = wb.create_sheet("Petrol 1")
        petrol["B2"] = "Dip"
        for i in range(50):
            petrol.cell(row=i + 3, column=2, value=i)
            petrol.cell(row=i + 3, column=3, value=i * 70)

        # An unrelated diesel sheet should not be picked for TANK-PETROL-1.
        d3 = wb.create_sheet(" Diesel 3")
        d3["B2"] = "Dip"
        for i in range(80):
            d3.cell(row=i + 3, column=2, value=i)
            d3.cell(row=i + 3, column=3, value=i * 4)

    res = _upload(client, owner_headers, "TANK-PETROL-1", _xlsx_bytes(build))
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["sheet_used"] == "Petrol 1"          # name-matched, not the densest
    assert body["point_count"] == 50


def test_template_endpoint_returns_xlsx(client):
    res = client.get("/api/v1/settings/tank-calibration/template")
    assert res.status_code == 200
    assert res.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # xlsx is a zip — first 2 bytes are 'PK'.
    assert res.content[:2] == b"PK"
