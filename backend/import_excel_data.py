"""
Import historical data from Excel spreadsheet into the fuel management system.
Reads daily tank readings from the Excel file and imports them via the API.
"""

import openpyxl
from datetime import datetime
import requests
import json
from openpyxl.utils import get_column_letter

# Configuration
EXCEL_FILE = "Daily Station Stock Movement Reconciliation Luanshya December 2025.xlsx"
API_BASE_URL = "http://localhost:8000/api/v1"
HEADER_ROW = 3
DATA_START_ROW = 4

# Column mapping (Excel columns to our fields)
COLUMN_MAP = {
    'B': 'date',
    'C': 'shift',

    # Nozzle 1A
    'D': 'attendant_1a',
    'E': 'electronic_opening_1a',
    'F': 'electronic_closing_1a',
    'H': 'mechanical_opening_1a',
    'I': 'mechanical_closing_1a',

    # Nozzle 1B
    'K': 'attendant_1b',
    'L': 'electronic_opening_1b',
    'M': 'electronic_closing_1b',
    'O': 'mechanical_opening_1b',
    'P': 'mechanical_closing_1b',

    # Nozzle 2A
    'R': 'attendant_2a',
    'S': 'electronic_opening_2a',
    'T': 'electronic_closing_2a',
    'V': 'mechanical_opening_2a',
    'W': 'mechanical_closing_2a',

    # Nozzle 2B
    'Y': 'attendant_2b',
    'Z': 'electronic_opening_2b',
    'AA': 'electronic_closing_2b',
    'AC': 'mechanical_opening_2b',
    'AD': 'mechanical_closing_2b',

    # Tank readings
    'AF': 'opening_dip_cm',
    'AG': 'after_delivery_dip_cm',
    'AH': 'closing_dip_cm',
    'AI': 'opening_volume',
    'AJ': 'before_offload_volume',
    'AK': 'after_offload_volume',
    'AL': 'closing_volume',

    # Financial
    'AX': 'price_per_liter',
    'BD': 'actual_cash_banked',

    # Customer Allocations (Diesel only - Columns AR-BB)
    'AR': 'customer_drive_in_volume',
    'AS': 'customer_volcano_volume',
    'AT': 'customer_hammington_volume',
    'AU': 'customer_special3_volume',
    'AV': 'customer_special4_volume',
    'AY': 'customer_volcano_price',
    'AZ': 'customer_hammington_price',
    'BA': 'customer_special3_price',
    'BB': 'customer_special4_price',
}

# Customer ID mapping
CUSTOMER_IDS = {
    'drive_in': 'CUST-DRIVE-IN',
    'volcano': 'CUST-VOLCANO',
    'hammington': 'CUST-HAMMINGTON',
    'special3': 'CUST-SPECIAL3',
    'special4': 'CUST-SPECIAL4',
}

def get_cell_value(ws, row, col_letter):
    """Get cell value by column letter and row number."""
    col_idx = openpyxl.utils.column_index_from_string(col_letter)
    cell = ws.cell(row, col_idx)

    # Handle None values
    if cell.value is None:
        return None

    # Handle datetime objects (for dates)
    if isinstance(cell.value, datetime):
        return cell.value

    # Convert to string and strip whitespace
    return str(cell.value).strip()

def parse_row(ws, row, tank_id):
    """Parse a single row from the Excel sheet."""
    # Get date
    date_value = get_cell_value(ws, row, 'B')
    if not date_value:
        return None

    # Parse date
    try:
        if isinstance(date_value, datetime):
            date_formatted = date_value.strftime('%Y-%m-%d')
        elif isinstance(date_value, str):
            # Try to parse string date
            if date_value == 'None':
                return None
            date_obj = datetime.strptime(date_value, '%Y-%m-%d %H:%M:%S')
            date_formatted = date_obj.strftime('%Y-%m-%d')
        else:
            return None
    except Exception as e:
        print(f"  [WARN] Skipping row {row}: Invalid date '{date_value}' - {e}")
        return None

    # Get shift type
    shift_raw = get_cell_value(ws, row, 'C')
    shift_type = 'Day' if shift_raw and 'day' in shift_raw.lower() else 'Night'

    # Parse nozzle readings
    nozzles = []

    # Nozzle 1A (columns D-J)
    attendant_1a = get_cell_value(ws, row, 'D')
    if attendant_1a and attendant_1a != 'None':
        elec_open_1a = get_cell_value(ws, row, 'E')
        elec_close_1a = get_cell_value(ws, row, 'F')
        mech_open_1a = get_cell_value(ws, row, 'H')
        mech_close_1a = get_cell_value(ws, row, 'I')

        if elec_open_1a and elec_close_1a:
            elec_open = float(elec_open_1a)
            elec_close = float(elec_close_1a)
            mech_open = float(mech_open_1a) if mech_open_1a and mech_open_1a != 'None' else 0
            mech_close = float(mech_close_1a) if mech_close_1a and mech_close_1a != 'None' else 0

            nozzles.append({
                'nozzle_id': '1A',
                'attendant': attendant_1a,
                'electronic_opening': elec_open,
                'electronic_closing': elec_close,
                'electronic_movement': elec_close - elec_open,
                'mechanical_opening': mech_open,
                'mechanical_closing': mech_close,
                'mechanical_movement': mech_close - mech_open,
            })

    # Nozzle 1B (columns K-Q)
    attendant_1b = get_cell_value(ws, row, 'K')
    if attendant_1b and attendant_1b != 'None':
        elec_open_1b = get_cell_value(ws, row, 'L')
        elec_close_1b = get_cell_value(ws, row, 'M')
        mech_open_1b = get_cell_value(ws, row, 'O')
        mech_close_1b = get_cell_value(ws, row, 'P')

        if elec_open_1b and elec_close_1b:
            elec_open = float(elec_open_1b)
            elec_close = float(elec_close_1b)
            mech_open = float(mech_open_1b) if mech_open_1b and mech_open_1b != 'None' else 0
            mech_close = float(mech_close_1b) if mech_close_1b and mech_close_1b != 'None' else 0

            nozzles.append({
                'nozzle_id': '1B',
                'attendant': attendant_1b,
                'electronic_opening': elec_open,
                'electronic_closing': elec_close,
                'electronic_movement': elec_close - elec_open,
                'mechanical_opening': mech_open,
                'mechanical_closing': mech_close,
                'mechanical_movement': mech_close - mech_open,
            })

    # Nozzle 2A (columns R-X)
    attendant_2a = get_cell_value(ws, row, 'R')
    if attendant_2a and attendant_2a != 'None':
        elec_open_2a = get_cell_value(ws, row, 'S')
        elec_close_2a = get_cell_value(ws, row, 'T')
        mech_open_2a = get_cell_value(ws, row, 'V')
        mech_close_2a = get_cell_value(ws, row, 'W')

        if elec_open_2a and elec_close_2a:
            elec_open = float(elec_open_2a)
            elec_close = float(elec_close_2a)
            mech_open = float(mech_open_2a) if mech_open_2a and mech_open_2a != 'None' else 0
            mech_close = float(mech_close_2a) if mech_close_2a and mech_close_2a != 'None' else 0

            nozzles.append({
                'nozzle_id': '2A',
                'attendant': attendant_2a,
                'electronic_opening': elec_open,
                'electronic_closing': elec_close,
                'electronic_movement': elec_close - elec_open,
                'mechanical_opening': mech_open,
                'mechanical_closing': mech_close,
                'mechanical_movement': mech_close - mech_open,
            })

    # Nozzle 2B (columns Y-AE)
    attendant_2b = get_cell_value(ws, row, 'Y')
    if attendant_2b and attendant_2b != 'None':
        elec_open_2b = get_cell_value(ws, row, 'Z')
        elec_close_2b = get_cell_value(ws, row, 'AA')
        mech_open_2b = get_cell_value(ws, row, 'AC')
        mech_close_2b = get_cell_value(ws, row, 'AD')

        if elec_open_2b and elec_close_2b:
            elec_open = float(elec_open_2b)
            elec_close = float(elec_close_2b)
            mech_open = float(mech_open_2b) if mech_open_2b and mech_open_2b != 'None' else 0
            mech_close = float(mech_close_2b) if mech_close_2b and mech_close_2b != 'None' else 0

            nozzles.append({
                'nozzle_id': '2B',
                'attendant': attendant_2b,
                'electronic_opening': elec_open,
                'electronic_closing': elec_close,
                'electronic_movement': elec_close - elec_open,
                'mechanical_opening': mech_open,
                'mechanical_closing': mech_close,
                'mechanical_movement': mech_close - mech_open,
            })

    # Skip row if no nozzle readings
    if not nozzles:
        return None

    # Parse tank readings
    opening_dip = get_cell_value(ws, row, 'AF')
    closing_dip = get_cell_value(ws, row, 'AH')

    if not opening_dip or not closing_dip:
        print(f"  [WARN] Skipping row {row}: Missing tank dip readings")
        return None

    # Check for delivery
    after_delivery_dip = get_cell_value(ws, row, 'AG')
    before_offload = get_cell_value(ws, row, 'AJ')
    after_offload = get_cell_value(ws, row, 'AK')

    delivery_occurred = bool(after_delivery_dip and after_delivery_dip != 'None')

    # Parse financial data
    price = get_cell_value(ws, row, 'AX')
    actual_cash = get_cell_value(ws, row, 'BD')

    # Parse customer allocations (Diesel only - Columns AR-BB)
    customer_allocations = []
    if tank_id == 'TANK-DIESEL':
        # Get default diesel price
        diesel_price = float(price) if price and price != 'None' else 26.98

        # Drive-In Customers (Column AR)
        drive_in_vol = get_cell_value(ws, row, 'AR')
        if drive_in_vol and drive_in_vol != 'None' and float(drive_in_vol) > 0:
            volume = float(drive_in_vol)
            customer_allocations.append({
                'customer_id': CUSTOMER_IDS['drive_in'],
                'customer_name': 'Drive-In Customers',
                'volume': volume,
                'price_per_liter': diesel_price,
                'amount': volume * diesel_price
            })

        # Volcano (Column AS with price AY)
        volcano_vol = get_cell_value(ws, row, 'AS')
        if volcano_vol and volcano_vol != 'None' and float(volcano_vol) > 0:
            volume = float(volcano_vol)
            volcano_price_raw = get_cell_value(ws, row, 'AY')
            volcano_price = float(volcano_price_raw) if volcano_price_raw and volcano_price_raw != 'None' else diesel_price
            customer_allocations.append({
                'customer_id': CUSTOMER_IDS['volcano'],
                'customer_name': 'Volcano',
                'volume': volume,
                'price_per_liter': volcano_price,
                'amount': volume * volcano_price
            })

        # Hammington (Column AT with price AZ)
        hammington_vol = get_cell_value(ws, row, 'AT')
        if hammington_vol and hammington_vol != 'None' and float(hammington_vol) > 0:
            volume = float(hammington_vol)
            hammington_price_raw = get_cell_value(ws, row, 'AZ')
            hammington_price = float(hammington_price_raw) if hammington_price_raw and hammington_price_raw != 'None' else diesel_price
            customer_allocations.append({
                'customer_id': CUSTOMER_IDS['hammington'],
                'customer_name': 'Hammington',
                'volume': volume,
                'price_per_liter': hammington_price,
                'amount': volume * hammington_price
            })

        # Special Customer 3 (Column AU with price BA)
        special3_vol = get_cell_value(ws, row, 'AU')
        if special3_vol and special3_vol != 'None' and float(special3_vol) > 0:
            volume = float(special3_vol)
            special3_price_raw = get_cell_value(ws, row, 'BA')
            special3_price = float(special3_price_raw) if special3_price_raw and special3_price_raw != 'None' else diesel_price
            customer_allocations.append({
                'customer_id': CUSTOMER_IDS['special3'],
                'customer_name': 'Special Customer 3',
                'volume': volume,
                'price_per_liter': special3_price,
                'amount': volume * special3_price
            })

        # Special Customer 4 (Column AV with price BB)
        special4_vol = get_cell_value(ws, row, 'AV')
        if special4_vol and special4_vol != 'None' and float(special4_vol) > 0:
            volume = float(special4_vol)
            special4_price_raw = get_cell_value(ws, row, 'BB')
            special4_price = float(special4_price_raw) if special4_price_raw and special4_price_raw != 'None' else diesel_price
            customer_allocations.append({
                'customer_id': CUSTOMER_IDS['special4'],
                'customer_name': 'Special Customer 4',
                'volume': volume,
                'price_per_liter': special4_price,
                'amount': volume * special4_price
            })

    # Build reading object
    reading = {
        'tank_id': tank_id,
        'date': date_formatted,
        'shift_type': shift_type,

        # Tank dips
        'opening_dip_cm': float(opening_dip),
        'closing_dip_cm': float(closing_dip),
        'after_delivery_dip_cm': float(after_delivery_dip) if delivery_occurred else None,

        # Tank volumes
        'opening_volume': float(get_cell_value(ws, row, 'AI')) if get_cell_value(ws, row, 'AI') and get_cell_value(ws, row, 'AI') != 'None' else None,
        'closing_volume': float(get_cell_value(ws, row, 'AL')) if get_cell_value(ws, row, 'AL') and get_cell_value(ws, row, 'AL') != 'None' else None,

        # Nozzle readings
        'nozzle_readings': nozzles,

        # Delivery
        'delivery_occurred': delivery_occurred,
        'before_offload_volume': float(before_offload) if before_offload and before_offload != 'None' else None,
        'after_offload_volume': float(after_offload) if after_offload and after_offload != 'None' else None,
        'delivery_time': None,
        'supplier': None,
        'invoice_number': None,

        # Financial
        'price_per_liter': float(price) if price and price != 'None' else 29.92,
        'actual_cash_banked': float(actual_cash) if actual_cash and actual_cash != 'None' else None,

        # Customer allocations (Diesel only)
        'customer_allocations': customer_allocations,

        'recorded_by': 'O001',  # Owner user ID
        'notes': 'Imported from Excel'
    }

    return reading

def import_sheet(ws, sheet_name, tank_id, token):
    """Import data from a single sheet."""
    print(f"\n{'='*60}")
    print(f"Importing {sheet_name} data...")
    print(f"{'='*60}")

    max_row = ws.max_row
    imported_count = 0
    skipped_count = 0
    error_count = 0

    print(f"Processing rows {DATA_START_ROW} to {max_row}...")

    for row in range(DATA_START_ROW, max_row + 1):
        # Parse row
        reading = parse_row(ws, row, tank_id)

        if not reading:
            skipped_count += 1
            continue

        # Submit to API
        try:
            response = requests.post(
                f"{API_BASE_URL}/tank-readings/readings",
                headers={
                    'Content-Type': 'application/json',
                    'Authorization': f'Bearer {token}'
                },
                json=reading
            )

            if response.status_code == 200:
                imported_count += 1
                allocation_info = ""
                if reading.get('customer_allocations'):
                    alloc_count = len(reading['customer_allocations'])
                    total_alloc_vol = sum(a['volume'] for a in reading['customer_allocations'])
                    allocation_info = f" | {alloc_count} customers, {total_alloc_vol:.1f}L allocated"
                print(f"  [OK] Row {row}: {reading['date']} ({reading['shift_type']}){allocation_info} - Imported successfully")
            else:
                error_count += 1
                error_detail = response.json().get('detail', 'Unknown error')
                print(f"  [ERROR] Row {row}: {reading['date']} ({reading['shift_type']}) - Error: {error_detail}")

        except Exception as e:
            error_count += 1
            print(f"  [ERROR] Row {row}: Exception - {str(e)}")

    print(f"\n{sheet_name} Import Summary:")
    print(f"  [OK] Imported: {imported_count}")
    print(f"  [WARN] Skipped: {skipped_count}")
    print(f"  [ERROR] Errors: {error_count}")

def main():
    """Main import function."""
    print("="*60)
    print("FUEL MANAGEMENT SYSTEM - EXCEL DATA IMPORT")
    print("="*60)

    # Login to get token
    print("\n[AUTH] Authenticating...")
    login_response = requests.post(
        f"{API_BASE_URL}/auth/login",
        json={
            'username': 'owner1',
            'password': 'owner123'
        }
    )

    if login_response.status_code != 200:
        print("[ERROR] Authentication failed. Please check credentials.")
        return

    token = login_response.json()['access_token']
    print("[OK] Authentication successful")

    # Load workbook with data_only=True to get calculated formula values
    print(f"\n[LOAD] Loading Excel file: {EXCEL_FILE}")
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        print(f"[OK] Excel file loaded successfully")
        print(f"  Available sheets: {', '.join(wb.sheetnames)}")
    except Exception as e:
        print(f"[ERROR] Error loading Excel file: {e}")
        return

    # Import Diesel data
    if 'Diesel' in wb.sheetnames:
        diesel_ws = wb['Diesel']
        import_sheet(diesel_ws, 'Diesel', 'TANK-DIESEL', token)
    else:
        print("[WARN] Diesel sheet not found")

    # Import Petrol data
    if 'Petrol' in wb.sheetnames:
        petrol_ws = wb['Petrol']
        import_sheet(petrol_ws, 'Petrol', 'TANK-PETROL', token)
    else:
        print("[WARN] Petrol sheet not found")

    print("\n" + "="*60)
    print("[COMPLETE] Import completed!")
    print("="*60)

if __name__ == "__main__":
    main()
