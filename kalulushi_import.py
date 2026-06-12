#!/usr/bin/env python3
"""
Kalulushi (ST001) live data import / reset  --  Neon PostgreSQL (JSONB storage).

Brings the live ST001 station to the real go-live state with 100% accuracy:
  * reshapes 6 islands / 12 nozzles  ->  the real 3 islands / 6 nozzles
  * sets the 6 opening nozzle readings (previous closing = this opening)
  * loads the 3 dip->volume calibration tables from the DIP CHAT workbook
  * sets opening tank levels (Petrol from dip 49.8 -> 5071 L; diesel from owner volumes)
  * clears all test transactional data (shifts, readings, handovers, sales, dailies...)
  * keeps config (settings, accounts, audit log) and the OTHER station (ST002) untouched

Modes (safe by default):
  python kalulushi_import.py --dry-run      # default; reads + prints the full plan, writes NOTHING
  python kalulushi_import.py --backup        # write a timestamped JSON backup of ST001, nothing else
  python kalulushi_import.py --apply --yes   # auto-backup, then apply in one transaction, then verify

Requires:  DATABASE_URL env var (Neon),  pip install "psycopg[binary]" openpyxl
"""
import os
import sys
import json
import argparse
import datetime

STATION_ID = "ST001"
CAL_XLSX = r"C:\Users\Purchase Requisition\Downloads\DIP CHAT (1) (1).xlsx"

# ---------------------------------------------------------------------------
# TARGET STATE (verbatim from the owner)
# ---------------------------------------------------------------------------
# nozzle_id -> dict of the fields the importer sets on that nozzle
NOZZLE_TARGETS = {
    "ISL1-A": dict(island="ISL-001", label="1A", fuel="Petrol", abbrev="UNL",
                   tank="TANK-PETROL",   elec=127124.12,   mech=1589250),
    "ISL1-B": dict(island="ISL-001", label="1B", fuel="Petrol", abbrev="UNL",
                   tank="TANK-PETROL",   elec=262562.59,   mech=2808291),
    "ISL2-A": dict(island="ISL-002", label="2A", fuel="Petrol", abbrev="UNL",
                   tank="TANK-PETROL",   elec=288471.720,  mech=1564748),
    "ISL2-B": dict(island="ISL-002", label="2B", fuel="Diesel", abbrev="LSD",
                   tank="TANK-DIESEL-2", elec=174912.548,  mech=799458),
    "ISL3-A": dict(island="ISL-003", label="3A", fuel="Diesel", abbrev="LSD",
                   tank="TANK-DIESEL",   elec=2954042.593, mech=2959642),
    "ISL3-B": dict(island="ISL-003", label="3B", fuel="Diesel", abbrev="LSD",
                   tank="TANK-DIESEL",   elec=2953080.241, mech=2957708),
}
ISLANDS_KEEP = ["ISL-001", "ISL-002", "ISL-003"]
ISLANDS_DELETE = ["ISL-004", "ISL-005", "ISL-006"]
ISLAND_PRODUCT_TYPE = {"ISL-001": "Petrol", "ISL-002": "Mixed", "ISL-003": "Diesel"}

# Opening tank levels (litres).
#   Petrol  : dip 49.8 -> 5071 L   (matches Petrol 1 calibration exactly)
#   Diesel  : owner-stated volumes (dips do not reconcile with the table - "use what we have")
# Diesel tank assignment CONFIRMED by owner:
#   "Diesel 1" = 14,000 L tank (TANK-DIESEL-2), dip 24.2 -> 1820 L
#   "Diesel 2" = 30,000 L tank (TANK-DIESEL),   dip 83.6 -> 7433 L
TANK_LEVELS = {
    "TANK-PETROL":   dict(litres=5071.0, note="dip 49.8 -> 5071 (verified)"),
    "TANK-DIESEL":   dict(litres=7433.0, note='owner "Diesel 2" (30,000 L) dip 83.6 -> 7433 (confirmed)'),
    "TANK-DIESEL-2": dict(litres=1820.0, note='owner "Diesel 1" (14,000 L) dip 24.2 -> 1820 (confirmed)'),
}

# DIP CHAT sheet name -> tank id   (note the leading space on " Diesel 3")
CAL_MAP = {"Petrol 1": "TANK-PETROL", "Diesel 2": "TANK-DIESEL", " Diesel 3": "TANK-DIESEL-2"}
CAL_PROVISIONAL = {"TANK-DIESEL-2"}  # 14,000 L tank but the only Diesel-3 table maxes at 30,500 L

# transactional storage keys to empty (test data) -> empty container of the right type
STORAGE_CLEAR = {
    "shifts": dict, "readings": list, "lpg_sales": list, "credit_sales": list,
    "lubricant_sales": list, "accessories_sales": list, "delivery_history": list,
    "reconciliations_data": list, "lpg_daily_entries": dict, "lpg_accessories_daily": dict,
    "lubricant_daily_entries": dict, "scheduled_price_changes": list,
}
# per-station files to empty (test data)
FILES_CLEAR = {
    "attendant_handovers.json": dict, "attendant_readings.json": dict,
    "reconciliations.json": list, "safe_deposits.json": dict, "sales.json": list,
    "tank_deliveries.json": dict, "lpg_accessories_daily.json": dict,
    "lpg_daily_entries.json": dict, "lubricant_daily_entries.json": dict,
    "notifications.json": list,
}
# kept as-is: audit_log.json (audit trail).  tank_calibrations.json -> rebuilt from the workbook.

# ---------------------------------------------------------------------------
DB = os.getenv("DATABASE_URL")
if not DB:
    sys.exit("ERROR: set DATABASE_URL (Neon connection string) first.")
try:
    import psycopg
    from psycopg.types.json import Jsonb
except ImportError:
    sys.exit('ERROR: pip install "psycopg[binary]"')


def now_iso():
    return datetime.datetime.now().isoformat()


def as_obj(v):
    if isinstance(v, (dict, list)) or v is None:
        return v
    try:
        return json.loads(v)
    except Exception:
        return v


def read_calibration():
    """sheet -> {tank_id: {chart:{dip_str:vol}, tank_id, point_count, uploaded_at, uploaded_by}}"""
    try:
        import openpyxl
    except ImportError:
        sys.exit('ERROR: pip install openpyxl')
    if not os.path.exists(CAL_XLSX):
        sys.exit(f"ERROR: calibration workbook not found: {CAL_XLSX}")
    wb = openpyxl.load_workbook(CAL_XLSX, data_only=True, read_only=True)
    out = {}
    for sheet, tank in CAL_MAP.items():
        if sheet not in wb.sheetnames:
            sys.exit(f"ERROR: sheet {sheet!r} not in workbook ({wb.sheetnames})")
        chart = {}
        for row in wb[sheet].iter_rows(values_only=True):
            # layout: (blank, Dip, Volume)
            if len(row) < 3:
                continue
            d, v = row[1], row[2]
            if isinstance(d, (int, float)) and isinstance(v, (int, float)):
                chart[f"{float(d):.1f}"] = float(v)
        out[tank] = dict(chart=chart, tank_id=tank, point_count=len(chart),
                         uploaded_at=now_iso(), uploaded_by="import:kalulushi_go_live")
    return out


# ---------------------------------------------------------------------------
# DB access
# ---------------------------------------------------------------------------
def fetch_all(cur):
    st = cur.execute("SELECT data FROM stations WHERE station_id=%s", (STATION_ID,)).fetchone()
    sto = cur.execute("SELECT data FROM station_storage WHERE station_id=%s", (STATION_ID,)).fetchone()
    files = {fn: as_obj(d) for fn, d in cur.execute(
        "SELECT filename, data FROM station_files WHERE station_id=%s", (STATION_ID,)).fetchall()}
    return (as_obj(st[0]) if st else None,
            as_obj(sto[0]) if sto else None,
            files)


def backup(cur, path=None):
    station, storage, files = fetch_all(cur)
    stamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    path = path or f"backup_{STATION_ID}_{stamp}.json"
    with open(path, "w", encoding="utf-8") as f:
        json.dump(dict(station_id=STATION_ID, taken_at=now_iso(),
                       station=station, station_storage=storage, station_files=files),
                  f, indent=2, default=str)
    print(f"  backup written: {os.path.abspath(path)}")
    return path


# ---------------------------------------------------------------------------
# Build the new storage dict (pure - no writes)
# ---------------------------------------------------------------------------
def transform(storage, calibration):
    report = []
    islands = storage.get("islands", {})

    # 1. delete extra islands
    for iid in ISLANDS_DELETE:
        if iid in islands:
            report.append(f"DELETE island {iid} ({islands[iid].get('name')})")
            del islands[iid]

    # 2. reshape kept islands' nozzles + set opening readings
    for iid in ISLANDS_KEEP:
        isl = islands.get(iid)
        if not isl:
            report.append(f"!! MISSING island {iid} - cannot set its nozzles")
            continue
        isl["product_type"] = ISLAND_PRODUCT_TYPE.get(iid, isl.get("product_type"))
        for nz in isl.get("pump_station", {}).get("nozzles", []):
            t = NOZZLE_TARGETS.get(nz.get("nozzle_id"))
            if not t:
                report.append(f"!! nozzle {nz.get('nozzle_id')} on {iid} has no target")
                continue
            before = (nz.get("fuel_type"), nz.get("tank_id"),
                      nz.get("electronic_reading"), nz.get("mechanical_reading"))
            nz["fuel_type"] = t["fuel"]
            nz["fuel_type_abbrev"] = t["abbrev"]
            nz["tank_id"] = t["tank"]
            nz["display_label"] = t["label"]
            nz["electronic_reading"] = t["elec"]
            nz["mechanical_reading"] = t["mech"]
            report.append(
                f"NOZZLE {nz['nozzle_id']} ({iid}): {before[0]}/{before[1]} "
                f"elec={before[2]} mech={before[3]}  ->  {t['fuel']}/{t['tank']} "
                f"elec={t['elec']} mech={t['mech']}")

    # 3. tank levels
    tanks = storage.get("tanks", {})
    for tid, spec in TANK_LEVELS.items():
        tk = tanks.get(tid)
        if not tk:
            report.append(f"!! MISSING tank {tid}")
            continue
        cap = tk.get("capacity") or 0
        before = tk.get("current_level")
        tk["current_level"] = spec["litres"]
        tk["percentage"] = round(spec["litres"] / cap * 100, 2) if cap else 0
        tk["last_updated"] = now_iso()
        report.append(f"TANK {tid}: level {before} -> {spec['litres']} L "
                      f"({tk['percentage']}% of {cap})  [{spec['note']}]")

    # 4. clear test transactional storage keys
    for key, kind in STORAGE_CLEAR.items():
        if key in storage:
            n = len(storage[key]) if isinstance(storage[key], (dict, list)) else "?"
            storage[key] = kind()
            report.append(f"CLEAR storage[{key}] ({n} -> 0)")

    return storage, report


def verify(cur, calibration):
    """Read back and assert everything equals target. Returns (ok, lines)."""
    _, storage, files = fetch_all(cur)
    ok, lines = True, []
    islands = storage.get("islands", {})

    for iid in ISLANDS_DELETE:
        good = iid not in islands
        ok &= good
        lines.append(f"[{'OK' if good else 'FAIL'}] island {iid} removed")
    n_islands = len(islands)
    g = n_islands == 3
    ok &= g
    lines.append(f"[{'OK' if g else 'FAIL'}] island count = {n_islands} (expect 3)")

    for nzid, t in NOZZLE_TARGETS.items():
        found = None
        for isl in islands.values():
            for nz in isl.get("pump_station", {}).get("nozzles", []):
                if nz.get("nozzle_id") == nzid:
                    found = nz
        good = bool(found) and found.get("fuel_type") == t["fuel"] and \
            found.get("tank_id") == t["tank"] and \
            found.get("electronic_reading") == t["elec"] and \
            found.get("mechanical_reading") == t["mech"]
        ok &= good
        lines.append(f"[{'OK' if good else 'FAIL'}] nozzle {nzid} "
                     f"{found.get('fuel_type') if found else '?'}/{found.get('tank_id') if found else '?'} "
                     f"elec={found.get('electronic_reading') if found else '?'}")

    tanks = storage.get("tanks", {})
    for tid, spec in TANK_LEVELS.items():
        lvl = tanks.get(tid, {}).get("current_level")
        good = lvl == spec["litres"]
        ok &= good
        lines.append(f"[{'OK' if good else 'FAIL'}] tank {tid} level={lvl} (expect {spec['litres']})")

    cal = files.get("tank_calibrations.json", {}) or {}
    for tid, rec in calibration.items():
        got = (cal.get(tid) or {}).get("point_count")
        good = got == rec["point_count"]
        ok &= good
        flag = "  (PROVISIONAL)" if tid in CAL_PROVISIONAL else ""
        lines.append(f"[{'OK' if good else 'FAIL'}] calibration {tid} points={got} "
                     f"(expect {rec['point_count']}){flag}")

    for key in STORAGE_CLEAR:
        v = storage.get(key)
        good = (not v)
        ok &= good
        lines.append(f"[{'OK' if good else 'FAIL'}] storage[{key}] empty")
    for fn in FILES_CLEAR:
        v = files.get(fn)
        good = (v is None) or (not v)
        ok &= good
        lines.append(f"[{'OK' if good else 'FAIL'}] file {fn} empty")
    return ok, lines


# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--backup", action="store_true")
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--verify-only", action="store_true")
    ap.add_argument("--yes", action="store_true", help="confirm a live --apply")
    args = ap.parse_args()
    mode = "apply" if args.apply else "backup" if args.backup else "verify" if args.verify_only else "dry-run"

    conn = psycopg.connect(DB, connect_timeout=30, autocommit=False)
    cur = conn.cursor()
    calibration = read_calibration()

    print(f"=== Kalulushi import  station={STATION_ID}  mode={mode} ===\n")
    print("Calibration tables read from workbook:")
    for tid, rec in calibration.items():
        flag = "  <-- PROVISIONAL (14,000 L tank, table maxes higher)" if tid in CAL_PROVISIONAL else ""
        print(f"  {tid:16s} {rec['point_count']} points{flag}")
    print()

    if mode == "verify":
        ok, lines = verify(cur, calibration)
        print("\n".join(lines))
        print("\nRESULT:", "ALL OK" if ok else "MISMATCH(ES) FOUND")
        conn.close()
        return

    if mode == "backup":
        backup(cur)
        conn.close()
        return

    # dry-run or apply: compute the plan
    _, storage, files = fetch_all(cur)
    if storage is None:
        sys.exit(f"ERROR: no station_storage row for {STATION_ID}")
    new_storage, report = transform(storage, calibration)

    print("PLANNED CHANGES")
    print("-" * 70)
    for line in report:
        print("  " + line)
    print("\n  REPLACE tank_calibrations.json with 3 tables (see counts above)")
    print("  CLEAR test files:")
    for fn in FILES_CLEAR:
        if fn in files:
            v = files[fn]
            n = len(v) if isinstance(v, (dict, list)) else "?"
            print(f"      {fn} ({n} -> 0)")
    print("  KEEP: audit_log.json, settings, accounts, users; ST002 untouched.")
    print("-" * 70)

    if mode == "dry-run":
        print("\nDRY-RUN ONLY - nothing was written.")
        print("Review the diesel tank assignments above, then re-run with:  --apply --yes")
        conn.close()
        return

    # APPLY
    if not args.yes:
        sys.exit("\nRefusing to apply without --yes. (Re-run: --apply --yes)")
    print("\nAPPLYING (auto-backup first)...")
    backup(cur)

    cur.execute("UPDATE station_storage SET data=%s WHERE station_id=%s",
                (Jsonb(new_storage), STATION_ID))
    # calibration: tank_calibrations.json content = {tank_id: rec}; upsert
    cal_file = calibration
    if cur.execute("UPDATE station_files SET data=%s WHERE station_id=%s AND filename=%s",
                   (Jsonb(cal_file), STATION_ID, "tank_calibrations.json")).rowcount == 0:
        cur.execute("INSERT INTO station_files (station_id, filename, data) VALUES (%s,%s,%s)",
                    (STATION_ID, "tank_calibrations.json", Jsonb(cal_file)))
    # clear files
    for fn, kind in FILES_CLEAR.items():
        if fn in files:
            cur.execute("UPDATE station_files SET data=%s WHERE station_id=%s AND filename=%s",
                        (Jsonb(kind()), STATION_ID, fn))

    conn.commit()
    print("Committed. Verifying...\n")
    ok, lines = verify(cur, calibration)
    print("\n".join(lines))
    print("\nRESULT:", "ALL OK - 100% match" if ok else ">>> MISMATCH(ES) - investigate (backup saved) <<<")
    conn.close()


if __name__ == "__main__":
    main()
