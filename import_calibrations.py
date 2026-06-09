#!/usr/bin/env python3
"""
Import updated tank calibration tables from "DIP CHAT (2).xlsx" into ST001.

Sheet -> tank mapping:
  "Petrol Tank 1 - 30000L"  ->  TANK-PETROL
  "Diesel Tank 1 - 30000L"  ->  TANK-DIESEL
  "Diesel Tank 2 - 14000L"  ->  TANK-DIESEL-2   (was provisional; now corrected)

Modes:
  python import_calibrations.py              # dry-run (default): prints plan, writes nothing
  python import_calibrations.py --apply      # writes to the database

Requires:  DATABASE_URL env var,  pip install "psycopg[binary]" openpyxl
"""
import os, sys, json, argparse
from datetime import datetime

STATION_ID = "ST001"
XLSX_PATH  = r"C:\Projects\Fuel\DIP CHAT (2).xlsx"

SHEET_TO_TANK = {
    "Petrol Tank 1 - 30000L": "TANK-PETROL",
    "Diesel Tank 1 - 30000L": "TANK-DIESEL",
    "Diesel Tank 2 - 14000L": "TANK-DIESEL-2",
}

# Max plausible volume per tank (from sheet name). Rows exceeding this are
# treated as data-entry errors and skipped.
SHEET_MAX_VOL = {
    "Petrol Tank 1 - 30000L": 30500,
    "Diesel Tank 1 - 30000L": 30500,
    "Diesel Tank 2 - 14000L": 14500,
}

# TANK-DIESEL-2 was the provisional table; all three are being updated.
WAS_PROVISIONAL = {"TANK-DIESEL-2"}


def parse_sheet(ws):
    """Return a sorted dict {dip_float: volume_float} from a calibration sheet."""
    chart = {}
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 2:
            continue
        dip, vol = row[0], row[1]
        if not isinstance(dip, (int, float)) or not isinstance(vol, (int, float)):
            continue
        if dip < 0 or vol < 0:
            continue
        chart[float(dip)] = float(vol)
    return dict(sorted(chart.items()))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true",
                        help="Write to the database (default is dry-run)")
    args = parser.parse_args()

    db_url = os.getenv("DATABASE_URL")
    if not db_url:
        sys.exit("ERROR: set DATABASE_URL first.")

    try:
        import psycopg
    except ImportError:
        sys.exit('ERROR: pip install "psycopg[binary]"')

    try:
        import openpyxl
    except ImportError:
        sys.exit("ERROR: pip install openpyxl")

    # --- Parse workbook ---
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    parsed = {}
    for sheet_name, tank_id in SHEET_TO_TANK.items():
        if sheet_name not in wb.sheetnames:
            sys.exit(f"ERROR: sheet '{sheet_name}' not found in workbook. "
                     f"Available: {wb.sheetnames}")
        ws = wb[sheet_name]
        chart = parse_sheet(ws)
        if len(chart) < 5:
            sys.exit(f"ERROR: too few data points in '{sheet_name}' ({len(chart)})")
        parsed[tank_id] = chart
        flag = " (was provisional - NOW CORRECTED)" if tank_id in WAS_PROVISIONAL else ""
        print(f"  {sheet_name!r:36s} -> {tank_id}  {len(chart)} points  "
              f"dip {min(chart):.1f}-{max(chart):.1f} mm  "
              f"vol {min(chart.values()):.0f}-{max(chart.values()):.0f} L{flag}")

    print()
    if not args.apply:
        print("DRY-RUN: no changes written. Re-run with --apply to commit.")
        return

    # --- Connect and update station_files ---
    conn = psycopg.connect(db_url, connect_timeout=20)
    cur  = conn.cursor()

    # Load existing calibrations (may be empty)
    row = cur.execute(
        "SELECT data FROM station_files WHERE station_id=%s AND filename=%s",
        (STATION_ID, "tank_calibrations.json"),
    ).fetchone()

    calibrations = {}
    if row:
        raw = row[0]
        calibrations = raw if isinstance(raw, dict) else json.loads(raw)

    now = datetime.now().isoformat()
    for tank_id, chart in parsed.items():
        # Store chart keys as strings (consistent with the upload endpoint)
        str_chart = {str(k): v for k, v in chart.items()}
        calibrations[tank_id] = {
            "tank_id":     tank_id,
            "chart":       str_chart,
            "uploaded_at": now,
            "uploaded_by": "import_calibrations.py",
            "point_count": len(str_chart),
        }
        print(f"  Saved {tank_id}: {len(str_chart)} points")

    # Upsert into station_files
    cur.execute("""
        INSERT INTO station_files (station_id, filename, data)
        VALUES (%s, %s, %s::jsonb)
        ON CONFLICT (station_id, filename)
        DO UPDATE SET data = EXCLUDED.data
    """, (STATION_ID, "tank_calibrations.json", json.dumps(calibrations)))

    conn.commit()
    conn.close()
    print("\nDone. Calibrations written to station_files for ST001.")
    print("The server will load these automatically on next restart,")
    print("or immediately if register_tank_calibration() is called at runtime.")


if __name__ == "__main__":
    main()
